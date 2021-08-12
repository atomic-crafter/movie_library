import tmdbsimple as tmdb
import requests
import tkinter
import sqlite3
from tkinter import ttk
import openpyxl
from openpyxl import Workbook

with open("DVD.txt") as movie_file:
	movies = movie_file.read().splitlines()



movie_db = sqlite3.connect("movie_database.db")
movie_db_cursor = movie_db.cursor()

def create_data(movie_db, movie_db_cursor,task):
	sql = """ INSERT INTO Movies_id(Movie_name,ID)
			VALUES(?,?)"""
	movie_db_cursor.execute(sql,task)
	movie_db.commit()
	return movie_db_cursor.lastrowid

def create_genre(movie_db, movie_db_cursor, task):
	sql = """ INSERT INTO Movies_genres(ID,Genres)
			VALUES(?,?)"""
	movie_db_cursor.execute(sql,task)
	movie_db.commit()
	return movie_db_cursor.lastrowid


def create_request(movie_db, movie_db_cursor, task):
	movie_db_cursor = movie_db.cursor()

	sql = """ INSERT INTO Old_requests(Title_requested)
			VALUES(?)"""
	movie_db_cursor.execute(sql,[task])
	movie_db.commit()
	#print("All clear cap'tain")
	return movie_db_cursor.lastrowid





tmdb.API_KEY = "INSERT YOUR KEY HERE"
tmdb.REQUESTS_SESSION = requests.Session()
search = tmdb.Search()
movie = tmdb.Movies(300)
response = movie.info()
#print(movie.title)


def get_ids(movie_db_cursor,slow):
	movie_db_cursor.execute("SELECT id FROM Movies_id")
	rows = movie_db_cursor.fetchall() #.replace(")","")  bug here
	


	for i in range(len(rows)):

		rows[i] = rows[i][0]

		"""
		rows[i].replace(")","")
		rows[i].replace("(","")
		rows[i].replace()
		
		"""

	return rows

def get_all_data():

	global id_names_dict,id_genres_dict,ids_list,movie_name_list,g_ids_list,genres_list
	
	movie_db_cursor = movie_db.cursor()
	movie_db_cursor.execute("SELECT id FROM Movies_id")
	ids_list = movie_db_cursor.fetchall() #.replace(")","")  bug here
	#print(ids_list)


	for i in range(len(ids_list)):

		ids_list[i] = ids_list[i][0]

		"""
		rows[i].replace(")","")
		rows[i].replace("(","")
		rows[i].replace()
		
		"""
	
	movie_db_cursor.execute("SELECT Movie_name FROM Movies_id")
	movie_name_list = movie_db_cursor.fetchall() #.replace(")","")  bug here
	#print(movie_name_list)


	for i in range(len(movie_name_list)):

		movie_name_list[i] = movie_name_list[i][0]

		"""
		rows[i].replace(")","")
		rows[i].replace("(","")
		rows[i].replace()
		
		"""
	id_names_dict = {}
	for i in range(len(ids_list)):
		id_names_dict[ids_list[i]] = movie_name_list[i]

	movie_db_cursor.execute("SELECT id FROM Movies_genres")
	g_ids_list = movie_db_cursor.fetchall() #.replace(")","")  bug here
	#print(g_ids_list)


	for i in range(len(g_ids_list)):

		g_ids_list[i] = g_ids_list[i][0]

		"""
		rows[i].replace(")","")
		rows[i].replace("(","")
		rows[i].replace()
		
		"""
	
	movie_db_cursor.execute("SELECT Genres FROM Movies_genres")
	genres_list = movie_db_cursor.fetchall() #.replace(")","")  bug here
	#print(genres_list)


	for i in range(len(genres_list)):

		genres_list[i] = genres_list[i][0]

		"""
		rows[i].replace(")","")
		rows[i].replace("(","")
		rows[i].replace()
		
		"""
	id_genres_dict = {}
	for i in range(len(g_ids_list)):
		id_genres_dict[g_ids_list[i]] = genres_list[i]

	#print(id_genres_dict)


	



def get_old_requests(movie_db_cursor):
	#print("I do not hope that")
	movie_db_cursor.execute("SELECT Title_requested FROM Old_requests")
	#print("I hope so...")
	rows = movie_db_cursor.fetchall() #.replace(")","")  bug here
	#print("okk")
	#print(rows)


	for i in range(len(rows)):

		rows[i] = rows[i][0]

		"""
		rows[i].replace(")","")
		rows[i].replace("(","")
		rows[i].replace()
		
		"""

	return rows


def get_genres(title, checked_ids, slow):
	#print(title)
	#create_request(movie_db, movie_db_cursor,title)
	#print("ok 1")
	checked_titles = get_old_requests(movie_db_cursor)
	if title not in checked_titles:
	

		#movie_db_cursor.execute("SELECT id FROM Movies_id")
		#rows = movie_db_cursor.fetchall() #.replace(")","")  bug here
		#print(rows)
		#print("ok 2")

		#print(title)
		response = search.movie(query=title,language = "fr")
		#print("so far so good")
		results_number = 0
		for s in search.results:
			if results_number == 0:
				print(s["title"],",id:",s["id"])
				id = s["id"]
				title = s["title"]
		results_number = 0
		#print(id)
		create_request(movie_db, movie_db_cursor,title)
	else:
		print("Already checked")

	if id not in checked_ids and id not in checked_titles:


		task = (title,id)

		create_data(movie_db,movie_db_cursor,task)

		movie_selected = tmdb.Movies(id)
		info_movie = movie_selected.info()



		#print(info_movie)
		#print(info_movie["genres"][0]["name"])


		movie_genres = list()

		for i in range(len(info_movie["genres"])):
			movie_genres.append(info_movie["genres"][i]["name"])
		#print(movie_genres)
		for i in range(len(movie_genres)):
			task = (id,movie_genres[i])
			create_genre(movie_db, movie_db_cursor, task)

		movie_genres_txt = ""

		for i in range(len(movie_genres)):
			movie_genres_txt += movie_genres[i]
			if len(movie_genres)-1 != i:
				movie_genres_txt +=", " 

		print("The genres of the movie you selected are:", movie_genres_txt)
	else:
		print("Deja dans la db")


gui = tkinter.Tk()
gui.geometry("300x100")
#def getEntry():
#	res = myEntry.get()
#	create_request(movie_db,movie_db_cursor,res)
#	get_genres(res)
def run_easy():
	checked_ids = get_ids(movie_db_cursor,False)
	for i in range(len(movies)):
		try:
			get_genres(movies[i], checked_ids,False)
			#create_request(movie_db,movie_db_cursor,movies[i])
		except:
			print("error")


def run_slow():
	checked_ids = get_ids(movie_db_cursor,True)
	for i in range(len(movies)):
		try:
			get_genres(movies[i], checked_ids,True)
			#create_request(movie_db,movie_db_cursor,movies[i])
		except:
			print("error")


def blit_infos():
	global id_names_dict,id_genres_dict,ids_list,movie_name_list,g_ids_list,genres_list
	global_list = list()

	for i in range(len(movie_name_list)):
		id = ids_list[i]
		movie_name = id_names_dict[id]
		try:
			genres = id_genres_dict[id]
		except:
			pass
		#genres = ""
		

		"""
		for j in range(len(temp_genres)):
			if j != 0:
				genres += " "
			print(temp_genres)
			print(type(temp_genres))
			genres += temp_genres[j]
		print(i)  """
		global_list.append((movie_name,id,genres))
		#print(global_list)





	root = tkinter.Tk()


	size_global_list = int(len(global_list))
	height = size_global_list
	width = 3
	frame = tkinter.Frame(root)
	frame.pack(pady = 20)
	scroll_table = ttk.Treeview(frame,columns = (1,2,3),show = "headings", height = 8)
	scroll_table.pack(side = tkinter.LEFT)

	wb = Workbook()

	ws = wb.active
	ws.title = "movie_library"

	ws.cell(row = 1, column = 1,value = "Id")
	ws.cell(row = 1, column = 2,value = "Title")
	ws.cell(row = 1, column = 3,value = "Genre")




	wb.save(filename = "movie_library.xlsx")



	scroll_table.heading(1, text = "Id")
	scroll_table.heading(2, text = "Title")
	scroll_table.heading(3, text = "Genre")

	scroll_bar = tkinter.Scrollbar(frame, orient = tkinter.VERTICAL)
	scroll_bar.pack(side = tkinter.RIGHT, fill = tkinter.Y)

	scroll_table.config(yscrollcommand = scroll_bar.set)
	scroll_bar.config(command = scroll_table.yview)

	for i in range(height): #Rows
		scroll_table.insert(parent="", index = i,values = (global_list[i][1],global_list[i][0],global_list[i][2]))

		ws.cell(row = i+2, column = 1,value = global_list[i][1])
		ws.cell(row = i+2, column = 2,value = global_list[i][0])
		ws.cell(row = i+2, column = 3,value = global_list[i][2])



	style = ttk.Style()
	style.theme_use("default")
	style.map("Treeview")

	tkinter.mainloop()
	wb.save(filename = "movie_library.xlsx")

def get_blit_data():
	get_all_data()
	blit_infos()


def get_entry(self):
	global myEntry

	res = myEntry.get()
	blit_search(res)
	return res

def search_movie_id():
	global myEntry
	search = tkinter.Tk()
	search.geometry("300x100")
	myEntry = tkinter.Entry(search)
	myEntry.pack()
	
	myEntry.bind("<Return>", get_entry)


def blit_search(res):
	global id_names_dict,id_genres_dict,ids_list,movie_name_list,g_ids_list,genres_list
	get_all_data()
	res = int(res)
	movie_name = id_names_dict[res]
	genre = id_genres_dict[res]
	root = tkinter.Tk()

	height = 1
	width = 3
	frame = tkinter.Frame(root)
	frame.pack(pady = 20)
	search_result = ttk.Treeview(frame,columns = (1,2,3),show = "headings", height = 8)
	search_result.pack(side = tkinter.LEFT)
	search_result.heading(1, text = "Id")
	search_result.heading(2, text = "Title")
	search_result.heading(3, text = "Genre")

	scroll_bar = tkinter.Scrollbar(frame, orient = tkinter.VERTICAL)
	scroll_bar.pack(side = tkinter.RIGHT, fill = tkinter.Y)

	search_result.config(yscrollcommand = scroll_bar.set)
	scroll_bar.config(command = search_result.yview)


	search_result.insert(parent="", index = 0,values = (res,movie_name,genre))







#movie_db_cursor.execute("SELECT id FROM Movies_id")

#rows = movie_db_cursor.fetchall().replace(")","")


#print(rows)

#for row in rows:
#	print(row)

#myEntry = tkinter.Entry(gui, width=40)
#myEntry.pack(pady=20)
btn = tkinter.Button(gui, height=1, width=15, text="Auto run programm", command=run_easy)
btn.pack()
"""
slow_btn = tkinter.Button(gui, height=1, width=27, text="Run programm and choose movies", command=run_slow)
slow_btn.pack()
"""
get_data_btn = tkinter.Button(gui, height=1, width=12, text="Show Data", command=get_blit_data)
get_data_btn.pack()

search_movie_btn = tkinter.Button(gui, height=1, width=15, text = "Search movie name",command = search_movie_id)
search_movie_btn.pack()


gui.mainloop()
